import csv
import openpyxl

def read_headers(input_filepath):
    with open(input_filepath, 'r') as text_file:
        headers = []
        for line in text_file:
            if line.strip() == '---':
                break
            headers.append(line.strip('|').split('|'))
        
        header_map = {}
        
        for header_row in headers:
            for i, header_cell in enumerate(header_row):
                if header_cell:
                    header_map[i] = header_cell
   
    return header_map

def read_data(input_filepath, header_map):
    with open(input_filepath, 'r') as text_file:
        data = []
        for line in text_file:
            data_row = line.strip('|').split('|')
            mapped_row = []
            for i in range(len(header_map)):
                if i in header_map and i < len(data_row) and data_row[i]:
                    mapped_row.append(data_row[i])
                else:
                    mapped_row.append('')

            data.append(mapped_row)
            
    return data

def convert_to_xlsx(header_map, data, output_filename):
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    # Write headers
    for i, header_cell in header_map.items():
        ws.cell(row=1, column=i+1).value = header_cell

    # Write data
    for i, data_row in enumerate(data):
        for j, data_cell in enumerate(data_row):
            ws.cell(row=i+2, column=j+1).value = data_cell

    wb.save(output_filename)

if __name__ == '__main__':
    input_filepath = input("Enter the input file path: ")
    output_filename = input_filepath.split('/')[-1].replace('.txt', '.xlsx')


    try:
        header_map = read_headers(input_filepath)
        print("Reading file..")
        data = read_data(input_filepath, header_map)
        print("Generating xlsx file..")
        convert_to_xlsx(header_map, data, output_filename)
        print("New file successful! File path: ", output_filename )
    except FileNotFoundError as error:
        print(error.args[0])
